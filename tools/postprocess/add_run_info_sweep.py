"""Add a ``run_info`` sheet to the sweep _paper xlsx (nodes / edges).

Mirrors the run_info convention used by the standard / adjacency-list reports
(config block + dataset-exact graph-size ranges + prompt counts), adapted for
the constraint-sweep layout (a node/edge axis instead of difficulties).

Graph-size ranges are read from the actual sample jsonls (``n_vertices`` /
``n_edges``), not parsed from prompts. Ranges are model-independent (same
seeds), so coloring ranges come from one coloring-special run and dc/sp ranges
from the qwen full sweep. Also fixes the summary's qwen ``model_pretrained``
label to Qwen/Qwen3.5-4B (the merge updated numbers but not that cell).

Run after merge_coloring_into_paper.py.
"""

from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _logs import find_sample_jsonls, iter_rows

REPO = Path(__file__).resolve().parents[2]
PAPER = REPO / "remote_results" / "_paper" / "_batch_reports"
RES = REPO / "remote_results"

MODELS = [
    "OpenGVLab/InternVL3_5-4B",
    "Qwen/Qwen3.5-4B",
    "google/gemma-4-E2B-it",
]
TASKS = ["coloring", "directed_connectivity", "shortest_path"]

TITLE_FONT = Font(bold=True, size=13, color="1F2A4A")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="305496")
SECTION_FONT = Font(bold=True, color="FFFFFF")
KEY_FONT = Font(bold=True)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

AXES = {
    "nodes": {
        "title": "Node-count sweep — run configuration",
        "axis_label": "node count",
        "values": list(range(3, 15)),  # 3..14
        "values_str": "3..14 (12 values)",
        "value_col": "nodes",
        "constraint_note": (
            "node count fixed to the axis value; edges follow from the graph"
        ),
    },
    "edges": {
        "title": "Edge-count sweep — run configuration",
        "axis_label": "edge count (target)",
        "values": [3, 5, 8, 12, 18, 25, 35],
        "values_str": "3, 5, 8, 12, 18, 25, 35 (7 values)",
        "value_col": "edges",
        "constraint_note": (
            "edge count rejection-sampled to the target by varying node count"
        ),
    },
}

SPV = 100  # samples per axis value


def _latest_paths(job: str) -> list[Path]:
    root = RES / job
    paths = find_sample_jsonls(root)
    tss = sorted(
        {m.group(1) for p in paths if (m := re.match(r"(\d{8}_\d{6})", p.name))}
    )
    if not tss:
        return []
    latest = tss[-1]
    return [p for p in paths if p.name.startswith(latest)]


def _ranges(job: str, want_tasks: set[str]) -> dict[tuple[str, int], list[int]]:
    """(task, constraint_value) -> [nv_min, nv_max, ne_min, ne_max, n_graphs]."""
    root = RES / job
    agg: dict[tuple[str, int], list[int]] = defaultdict(
        lambda: [10**9, -1, 10**9, -1, 0]
    )
    for r in iter_rows(_latest_paths(job), job_label=job, dataset_root=root):
        if r.base_task not in want_tasks or r.variant != "direct":
            continue
        a = agg[(r.base_task, r.constraint_value)]
        a[0] = min(a[0], r.n_vertices)
        a[1] = max(a[1], r.n_vertices)
        a[2] = min(a[2], r.n_edges)
        a[3] = max(a[3], r.n_edges)
        a[4] += 1
    return agg


def _config_rows(axis: str) -> list[tuple[str, str]]:
    a = AXES[axis]
    return [
        ("Models", ", ".join(MODELS)),
        ("Tasks", ", ".join(TASKS)),
        ("Sweep axis", f"{a['axis_label']} — values {a['values_str']}"),
        ("Samples per axis value",
         f"{SPV}  (→ {SPV} direct + {SPV} disguise prompts per value, per task)"),
        ("Coloring construction",
         "special-coloring: chromatic number planted uniformly over {2,3,4}"
         + ("; clamped to ≤ node count on the node axis" if axis == "nodes" else "")),
        ("conn / shortest_path construction",
         f"difficulty=medium presets; {a['constraint_note']}"),
        ("qwen35_4b model",
         "Qwen/Qwen3.5-4B — coloring + directed_connectivity + shortest_path re-run at this model"),
        ("gemma / internvl dc & shortest_path",
         "from the original sweep (same models); coloring re-run with special-χ"),
        ("Render settings",
         "label_style=numeric, node_color=#AED6F1, edge_style=straight"),
        ("Edge-count convention",
         "directed_connectivity = directed out-edges; coloring/shortest_path = undirected"),
        ("Graphs across models",
         "identical per (task, axis value) — same seeds; ranges below from one model"),
    ]


def build_run_info(axis: str, wb: openpyxl.Workbook) -> None:
    a = AXES[axis]
    # coloring ranges from a coloring-special run; dc/sp from the qwen full sweep
    color = _ranges(f"graph_bench_sweep_{axis}_coloring_internvl35_4b", {"coloring"})
    dcsp = _ranges(
        f"graph_bench_sweep_{axis}_qwen35_4b",
        {"directed_connectivity", "shortest_path"},
    )
    sizes = {**color, **dcsp}

    if "run_info" in wb.sheetnames:
        del wb["run_info"]
    ws = wb.create_sheet("run_info", 0)

    r = 1
    ws.cell(r, 1, a["title"]).font = TITLE_FONT
    r += 2
    for k, v in _config_rows(axis):
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 1).alignment = LEFT
        c = ws.cell(r, 2, v)
        c.alignment = LEFT
        r += 1
    r += 1

    # --- size range table ---
    ws.cell(r, 1, f"Graph-size range per task × {a['value_col']} value (dataset-exact)").font = KEY_FONT
    r += 1
    hdr = ["task", a["value_col"], "nodes_min", "nodes_max", "edges_min", "edges_max", "n_graphs"]
    for j, h in enumerate(hdr, 1):
        cell = ws.cell(r, j, h)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = CENTER
    r += 1
    for task in TASKS:
        for v in a["values"]:
            key = (task, v)
            if key not in sizes:
                continue
            nvmn, nvmx, nemn, nemx, ng = sizes[key]
            row = [task, v, nvmn, nvmx, nemn, nemx, ng]
            for j, val in enumerate(row, 1):
                cell = ws.cell(r, j, val)
                cell.alignment = LEFT if j == 1 else CENTER
            r += 1
    r += 1

    # --- prompt counts ---
    n_values = len(a["values"])
    per = SPV * n_values  # direct prompts per (model, task)
    ws.cell(r, 1, "Total prompts per model × task (direct + disguise, summed over axis values)").font = KEY_FONT
    r += 1
    for j, h in enumerate(["model", "task", "direct", "disguise", "total"], 1):
        cell = ws.cell(r, j, h)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = CENTER
    r += 1
    grand = 0
    for m in MODELS:
        for task in TASKS:
            row = [m, task, per, per, 2 * per]
            grand += 2 * per
            for j, val in enumerate(row, 1):
                ws.cell(r, j, val).alignment = LEFT if j <= 2 else CENTER
            r += 1
    r += 1
    ws.cell(r, 1, "Grand total prompts (all models × tasks × axis values)").font = KEY_FONT
    ws.cell(r, 5, grand).alignment = CENTER
    ws.cell(r, 5).font = KEY_FONT

    # widths
    for col, w in {"A": 42, "B": 60, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def fix_qwen_label(wb: openpyxl.Workbook) -> None:
    s = wb["summary"]
    hdr = [c.value for c in s[1]]
    mi = hdr.index("model") + 1
    pi = hdr.index("model_pretrained") + 1
    for row in range(2, s.max_row + 1):
        if s.cell(row, mi).value == "qwen35_4b":
            s.cell(row, pi).value = "Qwen/Qwen3.5-4B"


def main() -> int:
    axes = [a for a in ("nodes", "edges") if a in sys.argv] or ["nodes", "edges"]
    for axis in axes:
        path = PAPER / f"ablation_sweep_{axis}_latest.xlsx"
        wb = openpyxl.load_workbook(path)
        build_run_info(axis, wb)
        fix_qwen_label(wb)
        wb.save(path)
        print(f"[{axis}] added run_info + fixed qwen label -> {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

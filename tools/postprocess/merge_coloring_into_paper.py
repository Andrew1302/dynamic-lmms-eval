"""In-place merge of the coloring-fix sweep re-run into the _paper sweep xlsx.

Updates ``remote_results/_paper/_batch_reports/ablation_sweep_{nodes,edges}_latest.xlsx``
by replacing, for each axis:

* **coloring** rows (all 3 models) with the balanced-χ special-coloring run, and
* **qwen35_4b directed_connectivity / shortest_path** rows with the fresh
  Qwen3.5-4B full sweep (the _paper file had those from the old Qwen3-VL-4B).

gemma/internvl directed_connectivity/shortest_path rows are left untouched
(their model never changed, and their jsonl isn't available locally anyway).

The update is value-only and in place, so all cell formatting is preserved.
The summary sheet's per-task and overall accuracies are recomputed.

Run AFTER the batch ``sweep_coloring_rerun.txt`` has fetched results into
``remote_results/graph_bench_sweep_*``.
"""

from __future__ import annotations

import os
import shutil
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from batch_report import collect_sweep_job  # reuse the exact per-value + paired logic

REPO = Path(__file__).resolve().parents[2]
RESULTS = REPO / "remote_results"
PAPER = RESULTS / "_paper" / "_batch_reports"
MODELS = ["gemma4_e2b", "internvl35_4b", "qwen35_4b"]
TASKS3 = ["coloring", "directed_connectivity", "shortest_path"]


def _counts(long_rows, base, variant):
    """(correct, n) summed over a task/variant's per-value rows."""
    c = n = 0
    for r in long_rows:
        if r["base_task"] == base and r["variant"] == variant:
            n += r["n"]
            c += round(r["accuracy"] * r["n"])
    return c, n


def _per_value(long_rows, base):
    """{(variant, x): (accuracy, n)} for one base task."""
    out = {}
    for r in long_rows:
        if r["base_task"] == base:
            out[(r["variant"], r["x"])] = (r["accuracy"], r["n"])
    return out


def _collect(job_name):
    rdir = RESULTS / job_name
    if not (rdir / "logs").is_dir():
        raise SystemExit(f"missing results for {job_name} under {rdir}")
    _, long_rows, ts = collect_sweep_job(
        job_name=job_name, results_dir=rdir, timestamp_override=None
    )
    return long_rows, ts


def _acc_cells_map(ws):
    """{(base, variant, x): row_idx} for a per-model long-form sheet."""
    m = {}
    for r in range(2, ws.max_row + 1):
        base = ws.cell(r, 1).value
        variant = ws.cell(r, 2).value
        x = ws.cell(r, 3).value
        m[(base, variant, x)] = r
    return m


def _pivot_map(ws):
    """{(base, variant, x): row_idx} and model->col for the sweep pivot."""
    hdr = [c.value for c in ws[1]]
    model_col = {hdr[i]: i + 1 for i in range(4, len(hdr))}
    rows = {}
    for r in range(2, ws.max_row + 1):
        rows[(ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)] = r
    return rows, model_col


def merge_axis(axis: str, dry_run: bool = False) -> None:
    path = PAPER / f"ablation_sweep_{axis}_latest.xlsx"
    if not path.is_file():
        raise SystemExit(f"target not found: {path}")
    wb = openpyxl.load_workbook(path)

    # --- gather new data -----------------------------------------------------
    # coloring (all models) from the special-coloring jobs
    color_long = {}
    for m in MODELS:
        long_rows, ts = _collect(f"graph_bench_sweep_{axis}_coloring_{m}")
        color_long[m] = long_rows
        print(f"[{axis}] coloring {m}: ts={ts}")
    # qwen dc/sp from the fresh full Qwen3.5 sweep
    qwen_long, qts = _collect(f"graph_bench_sweep_{axis}_qwen35_4b")
    print(f"[{axis}] qwen full sweep: ts={qts}")

    # --- 1) per-model long-form tabs ----------------------------------------
    for m in MODELS:
        ws = wb[f"{axis}_{m}"]
        cells = _acc_cells_map(ws)
        # coloring rows (all models)
        for (variant, x), (acc, n) in _per_value(color_long[m], "coloring").items():
            ridx = cells[("coloring", variant, x)]
            ws.cell(ridx, 4).value = n
            ws.cell(ridx, 5).value = acc
        # qwen dc/sp rows
        if m == "qwen35_4b":
            for base in ("directed_connectivity", "shortest_path"):
                for (variant, x), (acc, n) in _per_value(qwen_long, base).items():
                    ridx = cells[(base, variant, x)]
                    ws.cell(ridx, 4).value = n
                    ws.cell(ridx, 5).value = acc

    # --- 2) sweep pivot ------------------------------------------------------
    sw = wb["sweep"]
    prows, mcol = _pivot_map(sw)
    # coloring: every model column
    for m in MODELS:
        col = mcol[m]
        for (variant, x), (acc, n) in _per_value(color_long[m], "coloring").items():
            sw.cell(prows[("coloring", variant, x)], col).value = acc
    # qwen dc/sp: only qwen column
    qcol = mcol["qwen35_4b"]
    for base in ("directed_connectivity", "shortest_path"):
        for (variant, x), (acc, n) in _per_value(qwen_long, base).items():
            sw.cell(prows[(base, variant, x)], qcol).value = acc

    # --- 3) summary ----------------------------------------------------------
    s = wb["summary"]
    hdr = [c.value for c in s[1]]
    col = {h: i + 1 for i, h in enumerate(hdr)}
    for r in range(2, s.max_row + 1):
        model = s.cell(r, col["model"]).value
        if model not in MODELS:
            continue
        # per-task accuracies: coloring from new; dc/sp new for qwen, else kept
        per_task = {}
        for variant in ("direct", "disguise", "paired"):
            cc, cn = _counts(color_long[model], "coloring", variant)
            per_task[("coloring", variant)] = (cc, cn)
            for base in ("directed_connectivity", "shortest_path"):
                if model == "qwen35_4b":
                    bc, bn = _counts(qwen_long, base, variant)
                else:  # keep existing accuracy; n assumed 100/value as before
                    existing = s.cell(r, col[f"{base}_{variant}_acc"]).value
                    # reconstruct counts from the per-model tab (unchanged rows)
                    bc, bn = _counts_from_tab(wb[f"{axis}_{model}"], base, variant)
                    _ = existing
                per_task[(base, variant)] = (bc, bn)
        # write per-task acc + recompute overall (sample-weighted across tasks)
        for variant in ("direct", "disguise", "paired"):
            for base in TASKS3:
                bc, bn = per_task[(base, variant)]
                s.cell(r, col[f"{base}_{variant}_acc"]).value = round(bc / bn, 4) if bn else ""
            tot_c = sum(per_task[(b, variant)][0] for b in TASKS3)
            tot_n = sum(per_task[(b, variant)][1] for b in TASKS3)
            s.cell(r, col[f"overall_{variant}_acc"]).value = round(tot_c / tot_n, 4) if tot_n else ""

    if dry_run:
        print(f"[{axis}] dry-run OK (not saved)")
        return
    bak = path.with_suffix(".xlsx.bak")
    shutil.copy2(path, bak)
    wb.save(path)
    print(f"[{axis}] saved {path}  (backup: {bak})")


def _counts_from_tab(ws, base, variant):
    """(correct, n) from an unchanged per-model long-form tab."""
    c = n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == base and ws.cell(r, 2).value == variant:
            nv = ws.cell(r, 4).value or 0
            av = ws.cell(r, 5).value or 0
            n += nv
            c += round(av * nv)
    return c, n


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    axes = [a for a in ("nodes", "edges") if a in sys.argv] or ["nodes", "edges"]
    for ax in axes:
        merge_axis(ax, dry_run=dry)

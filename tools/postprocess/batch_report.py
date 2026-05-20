"""Aggregate every job in a batch into a single Excel workbook.

Driven by ``remote_execution_scripts/07_batch_report.sh``. The wrapper
sources each job's ``.conf`` (bash) and hands us a small TSV listing
each job's ``results_dir``, ``COMPARE_PAIRS``, ``MODEL_PRETRAINED``;
output / batch-name / strictness / timestamp-override come in as CLI
flags.

Two layouts depending on the job kind:

* **Paired jobs** (``COMPARE_PAIRS`` non-empty) — uses
  ``compare_direct_disguise.build_pair_rows`` for the heavy lifting so the
  numbers match the per-job 06 xlsx exactly. Per-job tab carries the
  ``consolidated`` summary only (one row per task pair).

* **Sweep jobs** (no ``COMPARE_PAIRS``) — groups per-sample rows by
  ``(base_task, variant, constraint_value)`` and emits long-form rows on
  the per-job tab. The summary row only carries the overall direct /
  disguise / paired accuracy.

Jobs whose results haven't been fetched yet (no ``logs/`` subtree or no
matching jsonls) are recorded as ``NO_DATA`` on the summary sheet and
contribute no per-job tab. Pass ``--strict`` to abort on missing data
instead.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _excel import (
    CELL_BORDER,
    CENTER,
    HEADER_FILL,
    HEADER_FONT,
    LEFT,
    autosize,
    style_header_row,
    write_accuracy_cell,
)
from _logs import SAMPLES_RE, SampleRow, iter_rows
from compare_direct_disguise import (
    build_pair_rows,
    find_samples,
    pick_timestamp,
    load_jsonl,
)


# --------------------------------------------------------------------------- #
# Job-id parsing
# --------------------------------------------------------------------------- #

# Order matters: more specific prefixes first so "labels_letters" wins over a
# generic "labels" match.
_AXIS_PREFIXES: list[tuple[str, str, str]] = [
    # (prefix, axis, axis_value)
    ("graph_bench_standard_",                "standard",     ""),
    ("graph_bench_ablation_labels_letters_", "labels",       "letters"),
    ("graph_bench_ablation_labels_none_",    "labels",       "none"),
    ("graph_bench_ablation_color_",          "color",        ""),
    ("graph_bench_ablation_adjmatrix_",      "adjmatrix",    ""),
    ("graph_bench_ablation_thinking_",       "thinking",     ""),
    ("graph_bench_ablation_size_",           "model_size",   ""),
    ("graph_bench_sweep_nodes_",             "sweep_nodes",  ""),
    ("graph_bench_sweep_edges_",             "sweep_edges",  ""),
]


def parse_job_id(job_name: str) -> dict[str, str]:
    """Split a job name into ``axis``, ``axis_value``, ``model_short``.

    Falls back to ``axis="unknown"`` for non-matching names so the report
    still includes them (just unsorted/uncategorised).
    """
    for prefix, axis, axis_value in _AXIS_PREFIXES:
        if job_name.startswith(prefix):
            return {
                "axis": axis,
                "axis_value": axis_value,
                "model_short": job_name[len(prefix):],
            }
    return {"axis": "unknown", "axis_value": "", "model_short": job_name}


# --------------------------------------------------------------------------- #
# Per-job metric collection
# --------------------------------------------------------------------------- #

_PAIR_RE = re.compile(r"^(?P<label>[^:]+):(?P<direct>[^:]+):(?P<disguise>[^:]+)$")


def _parse_pair_spec(spec: str) -> tuple[str, str, str]:
    m = _PAIR_RE.match(spec)
    if not m:
        raise ValueError(f"bad COMPARE_PAIRS entry: {spec!r}")
    return m.group("label"), m.group("direct"), m.group("disguise")


def collect_paired_job(
    *,
    job_name: str,
    results_dir: Path,
    pairs: list[tuple[str, str, str]],
    timestamp_override: str | None,
) -> tuple[dict, list[dict], str]:
    """Return (summary_partial, consolidated_rows, picked_ts) for a paired job.

    ``summary_partial`` carries per-task and overall accuracy keys but no
    identity columns (axis/model/etc.) — the caller layers those on top.
    ``consolidated_rows`` mirrors compare_direct_disguise's ``main`` table.
    """
    inventory = find_samples(results_dir)
    referenced = {t for _, d, g in pairs for t in (d, g)}
    picked_ts = pick_timestamp(inventory, referenced, timestamp_override)
    by_task_ts = {(t, ts): p for t, entries in inventory.items() for ts, p in entries}

    consolidated: list[dict] = []
    per_task_acc: dict[str, dict[str, float | int]] = {}
    sum_direct_correct = sum_direct_total = 0
    sum_disguise_correct = sum_disguise_total = 0
    sum_paired_correct = sum_paired_total = 0

    for label, direct_task, disguise_task in pairs:
        d_path = by_task_ts[(direct_task, picked_ts)]
        g_path = by_task_ts[(disguise_task, picked_ts)]
        direct = load_jsonl(d_path)
        disguise = load_jsonl(g_path)
        _, totals = build_pair_rows(direct, disguise)

        n_direct = totals["n_direct"] or 1
        n_disguise = totals["n_disguise"] or 1
        n_paired = totals["n_paired"] or 1
        d_acc = totals["direct_correct"] / n_direct
        g_acc = totals["disguise_correct"] / n_disguise
        p_acc = totals["paired_correct"] / n_paired

        consolidated.append({
            "pair": label,
            "n_samples": totals["n_paired"],
            "direct_accuracy": round(d_acc, 4),
            "disguise_accuracy": round(g_acc, 4),
            "paired_accuracy": round(p_acc, 4),
        })
        per_task_acc[label] = {
            "direct": round(d_acc, 4),
            "disguise": round(g_acc, 4),
            "paired": round(p_acc, 4),
            "n": totals["n_paired"],
        }

        sum_direct_correct += totals["direct_correct"]
        sum_direct_total += totals["n_direct"]
        sum_disguise_correct += totals["disguise_correct"]
        sum_disguise_total += totals["n_disguise"]
        sum_paired_correct += totals["paired_correct"]
        sum_paired_total += totals["n_paired"]

    # n_samples is the per-pair count, not the cross-pair sum. Each pair-task
    # (coloring/directed_connectivity/shortest_path) evaluates the *same*
    # 5k graph instances independently, so summing across pairs would
    # triple-count the same samples. All pairs are expected to share the
    # same n_paired; we report the max and trust the per_task breakdown
    # to surface any divergence.
    n_per_pair = max((row["n_samples"] for row in consolidated), default=0)
    summary_partial = {
        "per_task": per_task_acc,
        "overall_direct_acc": round(sum_direct_correct / (sum_direct_total or 1), 4),
        "overall_disguise_acc": round(sum_disguise_correct / (sum_disguise_total or 1), 4),
        "overall_paired_acc": round(sum_paired_correct / (sum_paired_total or 1), 4),
        "n_samples": n_per_pair,
    }
    return summary_partial, consolidated, picked_ts


def collect_sweep_job(
    *,
    job_name: str,
    results_dir: Path,
    timestamp_override: str | None,
) -> tuple[dict, list[dict], str]:
    """Long-form rows for a sweep job + a coarse overall accuracy."""
    inventory = find_samples(results_dir)
    if not inventory:
        raise FileNotFoundError(f"no *_samples_*.jsonl under {results_dir}")
    all_tasks = set(inventory)
    picked_ts = pick_timestamp(inventory, all_tasks, timestamp_override)
    paths = [p for t, entries in inventory.items() for ts, p in entries if ts == picked_ts]

    # Group by (base_task, variant, constraint_value) → [rows]
    buckets: dict[tuple[str, str, int], list[int]] = defaultdict(list)
    total_direct = [0, 0]      # [correct, n]
    total_disguise = [0, 0]
    for row in iter_rows(paths, job_label=job_name):
        key = (row.base_task, row.variant, row.constraint_value)
        buckets[key].append(row.correct)
        if row.variant == "direct":
            total_direct[0] += row.correct
            total_direct[1] += 1
        elif row.variant == "disguise":
            total_disguise[0] += row.correct
            total_disguise[1] += 1

    long_rows: list[dict] = []
    for (base, variant, x), corrects in sorted(buckets.items()):
        n = len(corrects)
        long_rows.append({
            "base_task": base,
            "variant": variant,
            "x": x,
            "n": n,
            "accuracy": round(sum(corrects) / n, 4) if n else 0.0,
        })

    summary_partial = {
        "per_task": {},
        "overall_direct_acc": round(total_direct[0] / (total_direct[1] or 1), 4),
        "overall_disguise_acc": round(total_disguise[0] / (total_disguise[1] or 1), 4),
        "overall_paired_acc": "",  # paired accuracy is meaningless for sweeps
        "n_samples": total_direct[1] + total_disguise[1],
    }
    return summary_partial, long_rows, picked_ts


# --------------------------------------------------------------------------- #
# Workbook layout
# --------------------------------------------------------------------------- #

_BASE_TASKS_DEFAULT = ["coloring", "directed_connectivity", "shortest_path"]


def _collect_base_tasks(summary_rows: list[dict]) -> list[str]:
    """Stable ordering: known tasks first, then any extras in sorted order."""
    seen = set()
    for r in summary_rows:
        seen.update(r.get("per_task", {}).keys())
    extras = sorted(seen - set(_BASE_TASKS_DEFAULT))
    return [t for t in _BASE_TASKS_DEFAULT if t in seen] + extras


def _write_summary(ws, summary_rows: list[dict], base_tasks: list[str]) -> None:
    headers = [
        "job", "axis", "axis_value", "model", "model_pretrained",
        "samples_ts", "n_samples",
    ]
    per_task_headers: list[tuple[str, str, str]] = []  # (header, task, variant)
    for t in base_tasks:
        for variant in ("direct", "disguise", "paired"):
            h = f"{t}_{variant}_acc"
            headers.append(h)
            per_task_headers.append((h, t, variant))
    headers += ["overall_direct_acc", "overall_disguise_acc", "overall_paired_acc"]

    ws.append(headers)
    style_header_row(ws)

    accuracy_headers = {h for h, _, _ in per_task_headers} | {
        "overall_direct_acc", "overall_disguise_acc", "overall_paired_acc",
    }

    for row in summary_rows:
        per_task = row.get("per_task", {})
        excel_row = [
            row["job"],
            row["axis"],
            row["axis_value"],
            row["model"],
            row.get("model_pretrained", ""),
            row.get("samples_ts", ""),
            row.get("n_samples", ""),
        ]
        for _, t, variant in per_task_headers:
            v = per_task.get(t, {}).get(variant, "")
            excel_row.append(v)
        excel_row += [
            row.get("overall_direct_acc", ""),
            row.get("overall_disguise_acc", ""),
            row.get("overall_paired_acc", ""),
        ]
        ws.append(excel_row)
        r = ws.max_row
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=idx)
            cell.border = CELL_BORDER
            if h in accuracy_headers and isinstance(cell.value, (int, float)) and cell.value != "":
                write_accuracy_cell(cell, float(cell.value))
            elif h == "n_samples" and isinstance(cell.value, int):
                cell.number_format = "#,##0"
                cell.alignment = CENTER
            elif h == "job":
                cell.alignment = LEFT
                cell.font = Font(bold=True)
            else:
                cell.alignment = LEFT
    ws.freeze_panes = "C2"
    autosize(ws)


def _write_consolidated_tab(ws, consolidated: list[dict]) -> None:
    headers = ["pair", "n_samples", "direct_accuracy", "disguise_accuracy", "paired_accuracy"]
    ws.append(headers)
    style_header_row(ws)
    acc_cols = {"direct_accuracy", "disguise_accuracy", "paired_accuracy"}
    for row in consolidated:
        ws.append([row[h] for h in headers])
        r = ws.max_row
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=idx)
            cell.border = CELL_BORDER
            if h in acc_cols and isinstance(cell.value, (int, float)):
                write_accuracy_cell(cell, float(cell.value))
            elif h == "n_samples":
                cell.number_format = "#,##0"
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
                if h == "pair":
                    cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    autosize(ws)


def _write_sweep_tab(ws, long_rows: list[dict]) -> None:
    headers = ["base_task", "variant", "x", "n", "accuracy"]
    ws.append(headers)
    style_header_row(ws)
    for row in long_rows:
        ws.append([row[h] for h in headers])
        r = ws.max_row
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=idx)
            cell.border = CELL_BORDER
            if h == "accuracy" and isinstance(cell.value, (int, float)):
                write_accuracy_cell(cell, float(cell.value))
            elif h in {"x", "n"}:
                cell.number_format = "#,##0"
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
    ws.freeze_panes = "A2"
    autosize(ws)


def _safe_sheet_title(name: str, used: set[str]) -> str:
    # Strip the redundant `graph_bench_` prefix; otherwise the model suffix
    # gets cut off by Excel's 31-char limit.
    short = name
    for prefix in ("graph_bench_ablation_", "graph_bench_sweep_", "graph_bench_"):
        if short.startswith(prefix):
            short = short[len(prefix):]
            break
    title = short[:31] or "job"
    base = title
    suffix = 1
    while title in used:
        tag = f"_{suffix}"
        title = (base[: 31 - len(tag)] + tag)
        suffix += 1
    used.add(title)
    return title


# --------------------------------------------------------------------------- #
# Driver
# --------------------------------------------------------------------------- #


_MSYS_DRIVE_RE = re.compile(r"^/([a-zA-Z])/")


def _normalize_path(p: str) -> str:
    """Translate MSYS-style ``/c/Users/...`` paths to ``C:/Users/...`` on Windows.

    Git-bash on Windows produces ``/c/`` paths from ``pwd``; native Python's
    pathlib treats those as relative, so ``Path("/c/...").is_dir()`` is always
    False. Normalize here so the bash wrapper can stay portable.
    """
    if os.name == "nt":
        m = _MSYS_DRIVE_RE.match(p)
        if m:
            return f"{m.group(1).upper()}:/" + p[3:]
    return p


def _read_jobs_tsv(path: Path) -> list[dict]:
    """Parse the TSV emitted by 07_batch_report.sh.

    Columns: job_id, job_name, results_dir, model_pretrained, compare_pairs,
    constraint. The pairs field is '|'-joined. ``constraint`` is "" for
    non-sweep jobs, or "nodes"/"edges" for sweep jobs.
    """
    jobs: list[dict] = []
    with path.open("r", encoding="utf-8", newline="") as fh:
        for row in csv.reader(fh, delimiter="\t"):
            if not row or not row[0].strip():
                continue
            if len(row) < 6:
                row = row + [""] * (6 - len(row))
            job_id, job_name, results_dir, model_pretrained, pairs_joined, constraint = row[:6]
            pairs_raw = [p for p in pairs_joined.split("|") if p] if pairs_joined else []
            jobs.append({
                "job_id": job_id,
                "job_name": job_name,
                "results_dir": _normalize_path(results_dir),
                "model_pretrained": model_pretrained,
                "compare_pairs": pairs_raw,
                "constraint": constraint,
            })
    return jobs


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--jobs-tsv", type=Path, required=True,
                    help="Per-job metadata TSV produced by 07_batch_report.sh.")
    ap.add_argument("--output", type=Path, required=True,
                    help="Path to write the .xlsx workbook.")
    ap.add_argument("--latest-link", type=Path, default=None,
                    help="Optional convenience path refreshed to a copy of --output.")
    ap.add_argument("--batch-name", default="batch",
                    help="Label used in log messages.")
    ap.add_argument("--timestamp", default=None,
                    help="Pin a specific {YYYYMMDD_HHMMSS} sample timestamp for every job.")
    ap.add_argument("--strict", action="store_true",
                    help="Abort if any job is missing data (default: emit NO_DATA row).")
    args = ap.parse_args(argv)

    timestamp_override: str | None = args.timestamp or None
    strict: bool = args.strict
    output = Path(_normalize_path(str(args.output)))
    latest_link = Path(_normalize_path(str(args.latest_link))) if args.latest_link else None
    batch_name = args.batch_name

    summary_rows: list[dict] = []
    per_job_sheets: list[tuple[str, str, list[dict]]] = []  # (sheet_title, kind, rows)

    for job in _read_jobs_tsv(args.jobs_tsv):
        job_id = job["job_id"]
        job_name = job["job_name"]
        results_dir = Path(job["results_dir"])
        pairs = [_parse_pair_spec(s) for s in job["compare_pairs"]]
        model_pretrained = job["model_pretrained"]
        meta = parse_job_id(job_name)

        base_row = {
            "job": job_id,
            "axis": meta["axis"],
            "axis_value": meta["axis_value"],
            "model": meta["model_short"],
            "model_pretrained": model_pretrained,
            "samples_ts": "",
            "n_samples": "",
            "per_task": {},
            "overall_direct_acc": "",
            "overall_disguise_acc": "",
            "overall_paired_acc": "",
        }

        logs_dir = results_dir / "logs"
        if not logs_dir.is_dir():
            msg = f"[batch_report] {job_id}: no logs/ under {results_dir} -- NO_DATA"
            if strict:
                raise SystemExit(msg + " (--strict)")
            print(msg)
            base_row["samples_ts"] = "NO_DATA"
            summary_rows.append(base_row)
            continue

        try:
            if pairs:
                # Every job has COMPARE_PAIRS; this drives the summary row.
                summary_partial, consolidated, picked_ts = collect_paired_job(
                    job_name=job_name,
                    results_dir=results_dir,
                    pairs=pairs,
                    timestamp_override=timestamp_override,
                )
                # Per-job tab: sweep jobs get the long-form (x, n, accuracy)
                # so the curve information is preserved; everyone else gets
                # the same `consolidated` table that 06 writes.
                if job["constraint"]:
                    _, long_rows, _ = collect_sweep_job(
                        job_name=job_name,
                        results_dir=results_dir,
                        timestamp_override=picked_ts,
                    )
                    per_job_sheets.append((job_name, "sweep", long_rows))
                else:
                    per_job_sheets.append((job_name, "consolidated", consolidated))
            else:
                # No COMPARE_PAIRS — treat the whole job as a sweep / single
                # variant and let collect_sweep_job derive an overall accuracy.
                summary_partial, long_rows, picked_ts = collect_sweep_job(
                    job_name=job_name,
                    results_dir=results_dir,
                    timestamp_override=timestamp_override,
                )
                per_job_sheets.append((job_name, "sweep", long_rows))
        except (FileNotFoundError, SystemExit, KeyError) as exc:
            msg = f"[batch_report] {job_id}: NO_DATA ({exc})"
            if strict:
                raise SystemExit(msg + " (--strict)")
            print(msg)
            base_row["samples_ts"] = "NO_DATA"
            summary_rows.append(base_row)
            continue

        base_row.update(summary_partial)
        base_row["samples_ts"] = picked_ts
        summary_rows.append(base_row)
        print(f"[batch_report] {job_id}: ts={picked_ts} n={base_row['n_samples']}")

    # Sort: keep manifest order grouped by axis_value first, then model — within
    # the same axis_value, models are adjacent for at-a-glance comparison.
    summary_rows.sort(key=lambda r: (r["axis"], r["axis_value"], r["model"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    base_tasks = _collect_base_tasks(summary_rows)
    _write_summary(ws, summary_rows, base_tasks)

    used: set[str] = {"summary"}
    # Per-job tabs in the same order as the summary sheet.
    sheet_by_job = {name: (kind, rows) for name, kind, rows in per_job_sheets}
    for row in summary_rows:
        # Identify job_name back out of job_id (strip optional "graph_benchmark/" prefix).
        jid = row["job"]
        jname = jid.split("/")[-1]
        if jname not in sheet_by_job:
            continue
        kind, rows = sheet_by_job[jname]
        sheet = wb.create_sheet(title=_safe_sheet_title(jname, used))
        if kind == "consolidated":
            _write_consolidated_tab(sheet, rows)
        else:
            _write_sweep_tab(sheet, rows)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"[batch_report] wrote {output}")

    if latest_link is not None:
        latest_link.parent.mkdir(parents=True, exist_ok=True)
        if latest_link.exists() or latest_link.is_symlink():
            latest_link.unlink()
        # Plain copy — symlinks are flaky across WSL/Windows boundaries.
        shutil.copy2(output, latest_link)
        print(f"[batch_report] refreshed {latest_link}")

    return 0


if __name__ == "__main__":
    sys.exit(main())

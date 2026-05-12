"""Build a one-stop Excel summary of all ablation results.

Each ``--logs`` argument is a glob of fetched ``logs/`` directories.
You also supply a small JSON manifest mapping each results path to
its (model, axis, value) tuple — used to slot rows into the right
sheets. Example manifest:

    [
      {"results": "remote_results/graph_bench_standard_qwen3vl_4b/logs",
       "model": "qwen3vl_4b", "axis": "standard", "value": ""},
      {"results": "remote_results/graph_bench_ablation_labels_letters_qwen3vl_4b/logs",
       "model": "qwen3vl_4b", "axis": "labels", "value": "letters"},
      ...
    ]

Sheets produced:
  - ``standard`` — model × base_task accuracy (direct & disguise columns).
  - ``ablation_labels``, ``ablation_color``, ``ablation_adjmatrix``,
    ``ablation_thinking``, ``ablation_model_size`` — same shape pivoted
    by the ablation value.
  - ``sweep_nodes`` / ``sweep_edges`` — long-form (model, base_task,
    variant, x, accuracy). Plots live elsewhere; this sheet is the
    machine-readable source.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _excel import autosize, style_header_row, write_accuracy_cell
from _logs import SampleRow, find_sample_jsonls, iter_rows


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--manifest", required=True,
                   help="JSON file mapping results dirs to (model, axis, value).")
    p.add_argument("--output", default="analysis/summary.xlsx")
    return p.parse_args()


def _agg_accuracy(rows: list[SampleRow]) -> dict[tuple[str, str], tuple[int, int]]:
    """Group rows by (base_task, variant) → (correct, total)."""
    out: dict[tuple[str, str], list[int]] = defaultdict(lambda: [0, 0])
    for r in rows:
        out[(r.base_task, r.variant)][0] += r.correct
        out[(r.base_task, r.variant)][1] += 1
    return {k: (c, n) for k, (c, n) in out.items()}


def _add_pivot_sheet(wb: Workbook, sheet_name: str, by_value: dict) -> None:
    """``by_value`` is dict[ablation_value][model][(base, variant)] = (c, n)."""
    if not by_value:
        return
    ws = wb.create_sheet(sheet_name)
    values = sorted(by_value.keys())
    models = sorted({m for v in by_value.values() for m in v.keys()})
    pairs = sorted({k for v in by_value.values() for m in v.values() for k in m.keys()})

    headers = ["model", "value"] + [f"{b}/{var}" for b, var in pairs]
    ws.append(headers)
    style_header_row(ws)

    for model in models:
        for value in values:
            stats = by_value[value].get(model, {})
            row = [model, value]
            ws.append(row)
            r = ws.max_row
            for idx, key in enumerate(pairs, start=3):
                c, n = stats.get(key, (0, 0))
                cell = ws.cell(row=r, column=idx)
                if n:
                    write_accuracy_cell(cell, c / n)
                else:
                    cell.value = ""
    autosize(ws)


def _add_sweep_sheet(wb: Workbook, sheet_name: str, rows_by_x: dict) -> None:
    """``rows_by_x`` is dict[(model, base, variant)][x] = (c, n)."""
    if not rows_by_x:
        return
    ws = wb.create_sheet(sheet_name)
    ws.append(["model", "base_task", "variant", "x", "correct", "total", "accuracy"])
    style_header_row(ws)
    for (model, base, variant), by_x in sorted(rows_by_x.items()):
        for x, (c, n) in sorted(by_x.items()):
            ws.append([model, base, variant, x, c, n, ""])
            r = ws.max_row
            if n:
                write_accuracy_cell(ws.cell(row=r, column=7), c / n)
    autosize(ws)


def main() -> int:
    args = parse_args()
    manifest = json.loads(Path(args.manifest).read_text(encoding="utf-8"))

    # Aggregates:
    standard: dict[str, dict[tuple[str, str], tuple[int, int]]] = defaultdict(dict)
    pivots: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    sweeps: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0, 0])))

    for entry in manifest:
        rdir = Path(entry["results"])
        axis = entry.get("axis", "standard")
        model = entry["model"]
        value = entry.get("value", "")
        rows = list(iter_rows(find_sample_jsonls(rdir), job_label=model))
        if not rows:
            print(f"[warn] no sample rows under {rdir}")
            continue

        if axis == "standard":
            standard[model] = _agg_accuracy(rows)
        elif axis.startswith("sweep_"):
            xkey = "n_vertices" if axis == "sweep_nodes" else "n_edges"
            for r in rows:
                x = getattr(r, xkey)
                if x <= 0:
                    continue
                bucket = sweeps[axis][(model, r.base_task, r.variant)][x]
                bucket[0] += r.correct
                bucket[1] += 1
        else:
            agg = _agg_accuracy(rows)
            pivots[axis][value][model] = agg

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Standard sheet
    if standard:
        ws = wb.create_sheet("standard")
        pairs = sorted({k for stats in standard.values() for k in stats.keys()})
        headers = ["model"] + [f"{b}/{var}" for b, var in pairs]
        ws.append(headers)
        style_header_row(ws)
        for model, stats in sorted(standard.items()):
            ws.append([model])
            r = ws.max_row
            for idx, key in enumerate(pairs, start=2):
                c, n = stats.get(key, (0, 0))
                cell = ws.cell(row=r, column=idx)
                if n:
                    write_accuracy_cell(cell, c / n)
                else:
                    cell.value = ""
        autosize(ws)

    for axis, by_value in pivots.items():
        _add_pivot_sheet(wb, axis, by_value)

    for axis, rows_by_x in sweeps.items():
        # Convert nested defaultdict to plain dict-with-tuple-values
        clean: dict = {}
        for k, by_x in rows_by_x.items():
            clean[k] = {x: tuple(v) for x, v in by_x.items()}
        _add_sweep_sheet(wb, axis, clean)

    wb.save(args.output)
    print(f"[xlsx] wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

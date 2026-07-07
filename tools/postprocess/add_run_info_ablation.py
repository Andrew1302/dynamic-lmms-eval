"""Add a ``run_info`` sheet to a per-arm label/color ablation _paper xlsx.

Mirrors the run_info convention used by the standard / adjacency-list reports
(config block + dataset-exact graph-size ranges + prompt counts), for the
difficulty-separated label-style / node-color ablations (n=100/task/difficulty).

One arm per file:
    letters | none | color

Graph-size ranges are read from the actual sample jsonls (``n_vertices`` /
``n_edges`` in the per-row ``accuracy`` block), not parsed from prompts. Ranges
are model- and arm-independent (same seeds; label_style/node_color are render-
only knobs that don't change graph structure), so they're read from one model
of the arm's own runs.

Usage:
    python tools/postprocess/add_run_info_ablation.py <arm> <xlsx_path>
"""

from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _logs import find_sample_jsonls, iter_rows

REPO = Path(__file__).resolve().parents[2]
RES = REPO / "remote_results"

MODELS = [
    "OpenGVLab/InternVL3_5-4B",
    "Qwen/Qwen3.5-4B",
    "google/gemma-4-E2B-it",
]
TASKS = ["coloring", "directed_connectivity", "shortest_path"]
DIFFICULTIES = ["easy", "medium", "hard"]
SPD = 100  # samples per task per difficulty
RANGE_MODEL = "internvl35_4b"  # arbitrary; ranges are model-independent

# arm -> (title, job_base, ablated_row, render_row)
ARMS = {
    "letters": {
        "title": "Label-style ablation (letters) — run configuration",
        "job_base": "graph_bench_ablation_labels_letters",
        "ablated": ("Ablated setting",
                    "label_style = letters (nodes labelled A, B, C, …)"),
        "render": ("Render settings",
                   "label_style=letters, node_color=#AED6F1, edge_style=straight"),
    },
    "none": {
        "title": "Label-style ablation (no labels) — run configuration",
        "job_base": "graph_bench_ablation_labels_none",
        "ablated": ("Ablated setting",
                    "label_style = none (nodes drawn unlabelled)"),
        "render": ("Render settings",
                   "label_style=none, node_color=#AED6F1, edge_style=straight"),
    },
    "color": {
        "title": "Node-color ablation (#F1948A) — run configuration",
        "job_base": "graph_bench_ablation_color",
        "ablated": ("Ablated setting",
                    "node_color = #F1948A (salmon); baseline node_color = #AED6F1"),
        "render": ("Render settings",
                   "label_style=numeric, node_color=#F1948A, edge_style=straight"),
    },
}

TITLE_FONT = Font(bold=True, size=13, color="1F2A4A")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="305496")
SECTION_FONT = Font(bold=True, color="FFFFFF")
KEY_FONT = Font(bold=True)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def _latest_paths(job: str) -> list[Path]:
    root = RES / job
    paths = find_sample_jsonls(root)
    tss = sorted(
        {m.group(1) for p in paths if (m := re.match(r"(\d{8}_\d{6})", p.name))}
    )
    if not tss:
        return []
    latest = tss[-1]
    return [p for p in paths if p.name.startswith(latest)]


def _range_for(job: str, want_task: str) -> list[int] | None:
    """[nv_min, nv_max, ne_min, ne_max, n_graphs] over direct rows of want_task."""
    agg = [10**9, -1, 10**9, -1, 0]
    for r in iter_rows(_latest_paths(job), job_label=job):
        if r.base_task != want_task or r.variant != "direct":
            continue
        agg[0] = min(agg[0], r.n_vertices)
        agg[1] = max(agg[1], r.n_vertices)
        agg[2] = min(agg[2], r.n_edges)
        agg[3] = max(agg[3], r.n_edges)
        agg[4] += 1
    return agg if agg[4] else None


def _sizes(job_base: str) -> dict[tuple[str, str], list[int]]:
    """(task, difficulty) -> range. coloring from the *_coloring_* jobs;
    directed_connectivity + shortest_path from the conn+sp jobs."""
    out: dict[tuple[str, str], list[int]] = {}
    for diff in DIFFICULTIES:
        crange = _range_for(f"{job_base}_coloring_{diff}_{RANGE_MODEL}", "coloring")
        if crange:
            out[("coloring", diff)] = crange
        cs_job = f"{job_base}_{diff}_{RANGE_MODEL}"
        for task in ("directed_connectivity", "shortest_path"):
            rng = _range_for(cs_job, task)
            if rng:
                out[(task, diff)] = rng
    return out


def _config_rows(arm: str) -> list[tuple[str, str]]:
    cfg = ARMS[arm]
    return [
        ("Models", ", ".join(MODELS)),
        ("Tasks", ", ".join(TASKS)),
        ("Difficulties", "easy, medium, hard (pure per-difficulty; no per-task override)"),
        ("Generations per task per difficulty",
         f"{SPD}  (→ {SPD} direct + {SPD} disguise prompts)"),
        cfg["ablated"],
        ("Baseline for comparison",
         "standard run — label_style=numeric, node_color=#AED6F1"),
        ("Prompt augmentation", "none — image only (no adjacency list in prompt)"),
        ("Coloring construction",
         "special-coloring: chromatic number planted uniformly over {2,3,4}"),
        ("conn / shortest_path construction", "standard difficulty presets"),
        cfg["render"],
        ("Edge-count convention",
         "directed_connectivity = directed out-edges; coloring/shortest_path = undirected"),
        ("Graphs across models & arms",
         "identical per (task, difficulty) — same seeds; label/color are render-only; "
         "ranges below from one model"),
    ]


def build_run_info(arm: str, wb: openpyxl.Workbook) -> None:
    cfg = ARMS[arm]
    sizes = _sizes(cfg["job_base"])

    if "run_info" in wb.sheetnames:
        del wb["run_info"]
    ws = wb.create_sheet("run_info", 0)

    r = 1
    ws.cell(r, 1, cfg["title"]).font = TITLE_FONT
    r += 2
    for k, v in _config_rows(arm):
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, v).alignment = LEFT
        r += 1
    r += 1

    # --- size range table ---
    ws.cell(r, 1, "Graph-size range per task × difficulty (dataset-exact)").font = KEY_FONT
    r += 1
    hdr = ["task", "difficulty", "nodes_min", "nodes_max", "edges_min", "edges_max", "n_graphs"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r, j, h)
        c.fill = SECTION_FILL; c.font = SECTION_FONT; c.alignment = CENTER
    r += 1
    for task in TASKS:
        for diff in DIFFICULTIES:
            key = (task, diff)
            if key not in sizes:
                continue
            nvmn, nvmx, nemn, nemx, ng = sizes[key]
            for j, val in enumerate([task, diff, nvmn, nvmx, nemn, nemx, ng], 1):
                ws.cell(r, j, val).alignment = LEFT if j == 1 else CENTER
            r += 1
    r += 1

    # --- prompt counts ---
    per = SPD * len(DIFFICULTIES)  # direct prompts per (model, task) over difficulties
    ws.cell(r, 1, "Total prompts per model × task (direct + disguise, summed over difficulties)").font = KEY_FONT
    r += 1
    for j, h in enumerate(["model", "task", "direct", "disguise", "total"], 1):
        c = ws.cell(r, j, h)
        c.fill = SECTION_FILL; c.font = SECTION_FONT; c.alignment = CENTER
    r += 1
    grand = 0
    for m in MODELS:
        for task in TASKS:
            for j, val in enumerate([m, task, per, per, 2 * per], 1):
                ws.cell(r, j, val).alignment = LEFT if j <= 2 else CENTER
            grand += 2 * per
            r += 1
    r += 1
    ws.cell(r, 1, "Grand total prompts (all models × tasks × difficulties)").font = KEY_FONT
    ws.cell(r, 5, grand).font = KEY_FONT
    ws.cell(r, 5).alignment = CENTER

    for col, w in {"A": 42, "B": 66, "C": 11, "D": 11, "E": 11, "F": 11, "G": 11}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def main() -> int:
    if len(sys.argv) != 3 or sys.argv[1] not in ARMS:
        print(f"usage: {sys.argv[0]} <{'|'.join(ARMS)}> <xlsx_path>", file=sys.stderr)
        return 2
    arm, path = sys.argv[1], Path(sys.argv[2])
    wb = openpyxl.load_workbook(path)
    build_run_info(arm, wb)
    wb.save(path)
    print(f"[{arm}] added run_info -> {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

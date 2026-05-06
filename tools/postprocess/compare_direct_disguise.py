"""Compare lmms-eval `_direct` and `_disguise` task variants pair-by-pair.

Reads `{TIMESTAMP}_samples_<task>.jsonl` files produced by `lmms-eval
--log_samples` under a fetched `remote_results/<job>/logs/...` tree and
writes one Excel workbook with:

* a `main` sheet — one row per pair, with direct / disguise / paired
  accuracy (paired = both direct and disguise correct on the same doc_id);
* one sheet per pair — one row per doc_id with both responses, both
  correctness flags, and a final TOTAL row.

Trusts the per-sample `accuracy.score` already written by the task's
`process_results` hook (see lmms_eval/tasks/dynamic_graph_benchmark/utils.py)
so this script is benchmark-agnostic: any task pair that emits
`accuracy.score` works.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Excel-style "Good"/"Bad" semantics for at-a-glance pass/fail readability.
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True, color="1F2A4A")
THIN_SIDE = Side(border_style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

SAMPLES_RE = re.compile(r"^(?P<ts>\d{8}_\d{6})_samples_(?P<task>.+)\.jsonl$")


def parse_pair(spec: str) -> tuple[str, str, str]:
    parts = spec.split(":")
    if len(parts) != 3 or not all(parts):
        raise argparse.ArgumentTypeError(
            f"--pair must be LABEL:DIRECT_TASK:DISGUISE_TASK, got {spec!r}"
        )
    return parts[0], parts[1], parts[2]


def find_samples(results_dir: Path) -> dict[str, list[tuple[str, Path]]]:
    """Map task name → list of (timestamp, path) pairs found under results_dir."""
    out: dict[str, list[tuple[str, Path]]] = defaultdict(list)
    for path in results_dir.rglob("*_samples_*.jsonl"):
        m = SAMPLES_RE.match(path.name)
        if m:
            out[m.group("task")].append((m.group("ts"), path))
    for task in out:
        out[task].sort(key=lambda x: x[0])
    return out


def pick_timestamp(
    inventory: dict[str, list[tuple[str, Path]]],
    referenced_tasks: set[str],
    override: str | None,
) -> str:
    if override is not None:
        for task in referenced_tasks:
            tss = {ts for ts, _ in inventory.get(task, [])}
            if override not in tss:
                raise SystemExit(
                    f"--timestamp {override} not present for task {task!r}. "
                    f"Available: {sorted(tss) or 'none'}"
                )
        return override

    common: set[str] | None = None
    for task in referenced_tasks:
        tss = {ts for ts, _ in inventory.get(task, [])}
        if not tss:
            raise SystemExit(
                f"No *_samples_{task}.jsonl files found under the results dir."
            )
        common = tss if common is None else (common & tss)
    if not common:
        per_task = {t: sorted({ts for ts, _ in inventory.get(t, [])}) for t in referenced_tasks}
        raise SystemExit(
            "No single timestamp covers every referenced task. "
            f"Pin one with --timestamp. Per-task timestamps: {per_task}"
        )
    return max(common)


def load_jsonl(path: Path) -> dict[int, dict]:
    rows: dict[int, dict] = {}
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            rows[rec["doc_id"]] = rec
    return rows


def _resp(rec: dict | None) -> str:
    if rec is None:
        return ""
    val = rec.get("filtered_resps", "")
    if isinstance(val, list):
        return val[0] if val else ""
    return str(val)


def _correct(rec: dict | None) -> int | None:
    if rec is None:
        return None
    acc = rec.get("accuracy")
    if isinstance(acc, dict) and "score" in acc:
        return int(round(float(acc["score"])))
    return None


def build_pair_rows(direct: dict[int, dict], disguise: dict[int, dict]):
    ids = sorted(set(direct) | set(disguise))
    rows = []
    sum_direct = sum_disguise = sum_both = 0
    n_paired = 0
    for doc_id in ids:
        d = direct.get(doc_id)
        g = disguise.get(doc_id)
        d_corr = _correct(d)
        g_corr = _correct(g)
        target = ""
        if d is not None:
            target = str(d.get("target", ""))
        elif g is not None:
            target = str(g.get("target", ""))
        both = ""
        if d_corr is not None and g_corr is not None:
            both_int = int(d_corr == 1 and g_corr == 1)
            both = both_int
            sum_both += both_int
            n_paired += 1
        if d_corr is not None:
            sum_direct += d_corr
        if g_corr is not None:
            sum_disguise += g_corr
        rows.append({
            "doc_id": doc_id,
            "target": target,
            "direct_response": _resp(d),
            "direct_correct": d_corr if d_corr is not None else "",
            "disguise_response": _resp(g),
            "disguise_correct": g_corr if g_corr is not None else "",
            "both_correct": both,
        })
    rows.append({
        "doc_id": "TOTAL",
        "target": "",
        "direct_response": "",
        "direct_correct": sum_direct,
        "disguise_response": "",
        "disguise_correct": sum_disguise,
        "both_correct": sum_both,
    })
    return rows, {
        "n_direct": len(direct),
        "n_disguise": len(disguise),
        "n_paired": n_paired,
        "direct_correct": sum_direct,
        "disguise_correct": sum_disguise,
        "paired_correct": sum_both,
    }


def _autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 60)


def _style_header_row(sheet) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    sheet.row_dimensions[1].height = 22


def _write_consolidated(sheet, main_rows: list[dict]) -> None:
    headers = ["pair", "n_samples",
               "direct_accuracy", "disguise_accuracy", "paired_accuracy"]
    sheet.append(headers)
    _style_header_row(sheet)

    accuracy_cols = {"direct_accuracy", "disguise_accuracy", "paired_accuracy"}
    for row in main_rows:
        sheet.append([row[h] for h in headers])
        excel_row = sheet.max_row
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=excel_row, column=idx)
            cell.border = CELL_BORDER
            if header in accuracy_cols:
                cell.number_format = "0.00%"
                cell.alignment = CENTER
            elif header == "n_samples":
                cell.number_format = "#,##0"
                cell.alignment = CENTER
            elif header == "pair":
                cell.font = Font(bold=True)
                cell.alignment = LEFT
            else:
                cell.alignment = LEFT

    sheet.freeze_panes = "A2"
    _autosize(sheet)


def _write_pair_sheet(sheet, rows: list[dict]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    _style_header_row(sheet)

    correct_cols = {"direct_correct", "disguise_correct", "both_correct"}
    response_cols = {"direct_response": "direct_correct",
                     "disguise_response": "disguise_correct"}
    last_idx = len(rows) - 1

    for r_idx, row in enumerate(rows):
        is_total = (r_idx == last_idx)
        sheet.append([row[h] for h in headers])
        excel_row = sheet.max_row
        for c_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=excel_row, column=c_idx)
            cell.border = CELL_BORDER
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                if header in correct_cols:
                    cell.number_format = "#,##0"
                    cell.alignment = CENTER
                elif header == "doc_id":
                    cell.alignment = LEFT
                else:
                    cell.alignment = LEFT
                continue
            if header == "doc_id":
                cell.alignment = CENTER
                cell.number_format = "#,##0"
            elif header in correct_cols:
                cell.alignment = CENTER
                value = row.get(header)
                if value == 1:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif value == 0:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
            elif header in response_cols:
                cell.alignment = LEFT
                paired_correct = row.get(response_cols[header])
                if paired_correct == 1:
                    cell.font = GREEN_FONT
                elif paired_correct == 0:
                    cell.font = RED_FONT
            elif header == "target":
                cell.alignment = CENTER
                cell.font = Font(bold=True)
            else:
                cell.alignment = LEFT

    sheet.freeze_panes = "B2"
    _autosize(sheet)


def write_workbook(out_path: Path, main_rows: list[dict], pair_sheets: list[tuple[str, list[dict]]]) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "consolidated"
    if main_rows:
        _write_consolidated(ws, main_rows)

    used_titles: set[str] = {"consolidated"}
    for label, rows in pair_sheets:
        title = label[:31] or "pair"
        suffix = 1
        base = title
        while title in used_titles:
            tag = f"_{suffix}"
            title = (base[: 31 - len(tag)] + tag)
            suffix += 1
        used_titles.add(title)

        sheet = wb.create_sheet(title=title)
        _write_pair_sheet(sheet, rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--results-dir", type=Path, required=True,
                    help="Local results dir for one job (e.g. remote_results/<job>/).")
    ap.add_argument("--output", type=Path, required=True,
                    help="Path to write the .xlsx workbook.")
    ap.add_argument("--pair", action="append", required=True, type=parse_pair,
                    metavar="LABEL:DIRECT_TASK:DISGUISE_TASK",
                    help="Repeatable. Label plus the two lmms-eval task names to compare.")
    ap.add_argument("--timestamp", default=None,
                    help="Pin a specific {YYYYMMDD_HHMMSS} run. Default: latest covering all tasks.")
    args = ap.parse_args(argv)

    results_dir = args.results_dir.resolve()
    if not results_dir.is_dir():
        raise SystemExit(f"--results-dir not found or not a directory: {results_dir}")

    inventory = find_samples(results_dir)
    referenced = {t for _, d, g in args.pair for t in (d, g)}
    timestamp = pick_timestamp(inventory, referenced, args.timestamp)
    print(f"[compare] using timestamp {timestamp}")

    by_task_ts: dict[tuple[str, str], Path] = {
        (t, ts): p for t, entries in inventory.items() for ts, p in entries
    }

    main_rows: list[dict] = []
    pair_sheets: list[tuple[str, list[dict]]] = []
    for label, direct_task, disguise_task in args.pair:
        d_path = by_task_ts[(direct_task, timestamp)]
        g_path = by_task_ts[(disguise_task, timestamp)]
        print(f"[compare] {label}: direct={d_path.name} disguise={g_path.name}")

        direct = load_jsonl(d_path)
        disguise = load_jsonl(g_path)
        if set(direct) != set(disguise):
            sym_diff = set(direct).symmetric_difference(disguise)
            print(
                f"[compare]   warning: doc_id sets differ for {label} "
                f"({len(sym_diff)} doc_ids in only one variant); "
                "both_correct computed on intersection only.",
                file=sys.stderr,
            )

        rows, totals = build_pair_rows(direct, disguise)
        pair_sheets.append((label, rows))

        n = totals["n_paired"] or 1
        main_rows.append({
            "pair": label,
            "direct_task": direct_task,
            "disguise_task": disguise_task,
            "n_samples": totals["n_paired"],
            "direct_correct": totals["direct_correct"],
            "direct_accuracy": round(totals["direct_correct"] / (totals["n_direct"] or 1), 4),
            "disguise_correct": totals["disguise_correct"],
            "disguise_accuracy": round(totals["disguise_correct"] / (totals["n_disguise"] or 1), 4),
            "paired_correct": totals["paired_correct"],
            "paired_accuracy": round(totals["paired_correct"] / n, 4),
        })

    write_workbook(args.output, main_rows, pair_sheets)
    print(f"[compare] wrote {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

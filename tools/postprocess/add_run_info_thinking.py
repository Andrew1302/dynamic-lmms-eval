"""Add a ``run_info`` sheet to the thinking-vs-no-thinking ablation _paper xlsx.

Mirrors the run_info convention used by the other graph-benchmark reports
(config block + dataset-exact graph-size ranges + prompt counts), specialised
for the thinking ablation: 2 models × 3 difficulties × 2 arms (think / nothink),
n=100/task/difficulty. Qwen3.5-4B is excluded from this ablation (its no-think
arm goes verbose and truncates).

Graph structure is seed-identical across arms and models for a given
(task, difficulty) — thinking is a decode-time toggle, not a render/graph knob —
so size ranges are read from a single model's think-arm runs (n_vertices /
n_edges in each row's ``accuracy`` block), not parsed from prompts.

Usage:
    python tools/postprocess/add_run_info_thinking.py <xlsx_path>
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _logs import find_sample_jsonls, iter_rows

REPO = Path(__file__).resolve().parents[2]
RES = REPO / "remote_results"

MODELS = [
    "OpenGVLab/InternVL3_5-4B",
    "google/gemma-4-E2B-it",
    "Qwen/Qwen3.5-4B",
]
TASKS = ["coloring", "directed_connectivity", "shortest_path"]
DIFFICULTIES = ["easy", "medium", "hard"]
ARMS = ["think", "nothink"]
SPD = 100  # samples per task per difficulty
JOB_BASE = "graph_bench_think_think"   # think arm — ranges are arm/model-independent
RANGE_MODEL = "internvl35_4b"

TITLE = "Thinking vs. no-thinking ablation — run configuration"

TITLE_FONT = Font(bold=True, size=13, color="1F2A4A")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="305496")
SECTION_FONT = Font(bold=True, color="FFFFFF")
KEY_FONT = Font(bold=True)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def _latest_paths(job: str) -> list[Path]:
    root = RES / job
    paths = find_sample_jsonls(root)
    tss = sorted(
        {m.group(1) for p in paths if (m := re.match(r"(\d{8}_\d{6})", p.name))}
    )
    if not tss:
        return []
    latest = tss[-1]
    return [p for p in paths if p.name.startswith(latest)]


def _range_for(job: str, want_task: str) -> list[int] | None:
    """[nv_min, nv_max, ne_min, ne_max, n_graphs] over direct rows of want_task."""
    agg = [10**9, -1, 10**9, -1, 0]
    for r in iter_rows(_latest_paths(job), job_label=job):
        if r.base_task != want_task or r.variant != "direct":
            continue
        agg[0] = min(agg[0], r.n_vertices)
        agg[1] = max(agg[1], r.n_vertices)
        agg[2] = min(agg[2], r.n_edges)
        agg[3] = max(agg[3], r.n_edges)
        agg[4] += 1
    return agg if agg[4] else None


def _sizes() -> dict[tuple[str, str], list[int]]:
    """(task, difficulty) -> range. coloring from the *_coloring_* jobs;
    directed_connectivity + shortest_path from the conn+sp jobs."""
    out: dict[tuple[str, str], list[int]] = {}
    for diff in DIFFICULTIES:
        crange = _range_for(f"{JOB_BASE}_coloring_{diff}_{RANGE_MODEL}", "coloring")
        if crange:
            out[("coloring", diff)] = crange
        cs_job = f"{JOB_BASE}_{diff}_{RANGE_MODEL}"
        for task in ("directed_connectivity", "shortest_path"):
            rng = _range_for(cs_job, task)
            if rng:
                out[(task, diff)] = rng
    return out


def _config_rows() -> list[tuple[str, str]]:
    return [
        ("Ablation axis", "thinking = on (think) vs. off (nothink); single checkpoint per model"),
        ("Models", ", ".join(MODELS)),
        ("Tasks", ", ".join(TASKS)),
        ("Difficulties", "easy, medium, hard (pure per-difficulty; no per-task override)"),
        ("Generations per task per difficulty",
         f"{SPD}  (→ {SPD} direct + {SPD} disguise prompts) per arm"),
        ("InternVL3.5-4B — think mechanism",
         "vLLM, fp8-quantized; R1 system prompt (variant internvl_r1_v1: faithful "
         "OpenGVLab R1 + commit rule); reasoning in <think>…</think>; budget 12,288 tok, temp 0.6"),
        ("Gemma-4-E2B — think mechanism",
         "vLLM, bf16; enable_thinking chat-template flag; reasoning in special tokens "
         "<|channel>thought … <channel|> (skip_special_tokens=False); budget 12,288 tok, temp 0.6"),
        ("Qwen3.5-4B — think mechanism",
         "vLLM, fp8-quantized; enable_thinking chat-template flag; reasoning in <think>…</think>. "
         "NATIVE thinking-token budget = 12,288 (vllm reasoning_parser=qwen3 + reasoning_config): "
         "force-closes </think> at the budget so heavy over-deliberation never truncates the answer "
         "(~40%→0% lost). max_new_tokens = 12,288 + 1,024 answer allowance = 13,312; temp 0.6"),
        ("No-think arm",
         "reasoning disabled (no R1 prompt / enable_thinking=false). InternVL/Gemma: task-default "
         "max_new_tokens. Qwen3.5: a reasoning_prompt directive coerces a terse answer, "
         "max_new_tokens=1,024"),
        ("Answer extraction",
         "reasoning stripped via task reasoning_tags, then answer regex on post-reasoning text; "
         "verified answer-isolation & extract=score both 100% (verify_thinking.py)"),
        ("Prompt augmentation", "none — image only (no adjacency list in prompt)"),
        ("Coloring construction",
         "special-coloring: chromatic number planted uniformly over {2,3,4}"),
        ("conn / shortest_path construction", "standard difficulty presets"),
        ("Render settings", "label_style=numeric, node_color=#AED6F1, edge_style=straight"),
        ("Edge-count convention",
         "directed_connectivity = directed out-edges; coloring/shortest_path = undirected"),
        ("Graphs across arms & models",
         "identical per (task, difficulty) — same seeds; thinking is a decode-time toggle; "
         "ranges below from one model's think-arm runs"),
    ]


def build_run_info(wb: openpyxl.Workbook) -> None:
    sizes = _sizes()

    if "run_info" in wb.sheetnames:
        del wb["run_info"]
    ws = wb.create_sheet("run_info", 0)

    r = 1
    ws.cell(r, 1, TITLE).font = TITLE_FONT
    r += 2
    for k, v in _config_rows():
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, v).alignment = LEFT
        r += 1
    r += 1

    # --- size range table ---
    ws.cell(r, 1, "Graph-size range per task × difficulty (dataset-exact)").font = KEY_FONT
    r += 1
    hdr = ["task", "difficulty", "nodes_min", "nodes_max", "edges_min", "edges_max", "n_graphs"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r, j, h)
        c.fill = SECTION_FILL; c.font = SECTION_FONT; c.alignment = CENTER
    r += 1
    for task in TASKS:
        for diff in DIFFICULTIES:
            key = (task, diff)
            if key not in sizes:
                continue
            nvmn, nvmx, nemn, nemx, ng = sizes[key]
            for j, val in enumerate([task, diff, nvmn, nvmx, nemn, nemx, ng], 1):
                ws.cell(r, j, val).alignment = LEFT if j == 1 else CENTER
            r += 1
    r += 1

    # --- prompt counts ---
    # direct prompts per (model, arm, task) over difficulties
    per = SPD * len(DIFFICULTIES)
    ws.cell(r, 1, "Total prompts per model × arm × task (direct + disguise, summed over difficulties)").font = KEY_FONT
    r += 1
    for j, h in enumerate(["model", "arm", "task", "direct", "disguise", "total"], 1):
        c = ws.cell(r, j, h)
        c.fill = SECTION_FILL; c.font = SECTION_FONT; c.alignment = CENTER
    r += 1
    grand = 0
    for m in MODELS:
        for arm in ARMS:
            for task in TASKS:
                for j, val in enumerate([m, arm, task, per, per, 2 * per], 1):
                    ws.cell(r, j, val).alignment = LEFT if j <= 3 else CENTER
                grand += 2 * per
                r += 1
    r += 1
    ws.cell(r, 1, "Grand total prompts (all models × arms × tasks × difficulties)").font = KEY_FONT
    ws.cell(r, 6, grand).font = KEY_FONT
    ws.cell(r, 6).alignment = CENTER

    for col, w in {"A": 42, "B": 74, "C": 20, "D": 11, "E": 11, "F": 11, "G": 11}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def main() -> int:
    if len(sys.argv) != 2:
        print(f"usage: {sys.argv[0]} <xlsx_path>", file=sys.stderr)
        return 2
    path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(path)
    build_run_info(wb)
    wb.save(path)
    print(f"[thinking] added run_info -> {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

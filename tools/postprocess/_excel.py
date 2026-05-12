"""Shared openpyxl styling helpers for the ablation/summary workbooks.

Extracted from ``compare_direct_disguise.py`` so the new analysis scripts
(``plot_accuracy_vs_nodes.py``, ``build_summary_table.py``) and the
existing compare script share one style vocabulary.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True, color="1F2A4A")
THIN_SIDE = Side(border_style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    ws.row_dimensions[row].height = 22


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(
            (len(str(c.value)) for c in col if c.value is not None), default=8
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 60)


def write_accuracy_cell(cell, value: float) -> None:
    cell.value = value
    cell.number_format = "0.00%"
    cell.alignment = CENTER
    cell.border = CELL_BORDER
    # Light pass/fail colouring at extreme thresholds. Avoids noise on
    # mid-range values where colour would just be distracting.
    if value >= 0.9:
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT
    elif value < 0.3:
        cell.fill = RED_FILL
        cell.font = RED_FONT

"""Add a ``metrics`` sheet to the thinking-ablation _paper xlsx.

Documents the trust/quality metrics behind the accuracy numbers, straight in the
workbook so the paper file is self-contained. Reuses the PRODUCTION diagnosis
from ``verify_thinking`` (same strip_reasoning_tags + _normalize the scorer
uses), so nothing drifts.

Per (model, arm) it reports: n, reasoning-emitted %, truncated % (reasoning
never closed → answer lost), answer-isolated %, extract=score %, accuracy. Then
a per-job token-usage table (median / p90 / max generated tokens + truncation).

Crucially it uses only the **latest merged run per job** (newest timestamp among
the non-chunk sample files), matching what 07_batch_report.sh reports — so the
Gemma coloring reruns are used and stale earlier-timestamp samples in the same
folder are ignored (aggregating the whole folder blends runs and misleads).

Usage:
    python tools/postprocess/add_metrics_thinking.py <xlsx_path>
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))

# Reuse the production diagnosis + loaders (no logic copy = no drift).
import verify_thinking as vt  # noqa: E402
from _logs import SAMPLES_RE, find_sample_jsonls  # noqa: E402

REPO = _HERE.parents[1]
RES = REPO / "remote_results"

MODELS = ["internvl35_4b", "gemma4_e2b", "qwen35_4b"]
ARMS = ["nothink", "think"]

# The 24 ablation jobs (2 models x 3 diff x 2 arms x {coloring, conn+sp}).
DIFFS = ["easy", "medium", "hard"]

SECTION_FILL = PatternFill(fill_type="solid", fgColor="305496")
SECTION_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="1F2A4A")
KEY_FONT = Font(bold=True)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

_TS_RE = re.compile(r"^(\d{8}_\d{6})")


def _job_dirs() -> list[Path]:
    dirs = []
    for arm in ARMS:
        for d in DIFFS:
            for model in MODELS:
                for base in (f"graph_bench_think_{arm}_coloring_{d}_{model}",
                             f"graph_bench_think_{arm}_{d}_{model}"):
                    p = RES / base
                    if p.is_dir():
                        dirs.append(p)
    return dirs


def _latest_merged_jsonls(job_dir: Path) -> list[Path]:
    """Sample files of the newest run in this job dir. Prefer merged (non-chunk)
    files; fall back to chunk files only if no merged file exists. Restricting to
    one timestamp is what keeps reruns from blending with stale earlier samples."""
    paths = find_sample_jsonls(job_dir)
    merged = [p for p in paths if "chunks" not in p.parts]
    pool = merged or paths
    tss = sorted({m.group(1) for p in pool if (m := _TS_RE.match(p.name))})
    if not tss:
        return []
    latest = tss[-1]
    return [p for p in pool if p.name.startswith(latest)]


def _collect() -> list[vt.Diag]:
    tags = vt._load_reasoning_tags(None)
    diags: list[vt.Diag] = []
    for job_dir in _job_dirs():
        for p in _latest_merged_jsonls(job_dir):
            m = SAMPLES_RE.match(p.name)
            if not m:
                continue
            label = vt._job_label(p)
            task_full = m.group("task")
            with p.open("r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line:
                        diags.append(vt._diagnose(json.loads(line), task_full, label, tags))
    return diags, tags


def _pct(num: int, den: int) -> float:
    return round(100.0 * num / den, 1) if den else 0.0


def _pctl(vals: list[int], q: float) -> int:
    if not vals:
        return 0
    s = sorted(vals)
    return s[min(len(s) - 1, int(q * (len(s) - 1) + 0.5))]


def build_metrics(wb: openpyxl.Workbook) -> None:
    diags, tags = _collect()
    if not diags:
        raise SystemExit("no sample rows found under remote_results for the think jobs")

    if "metrics" in wb.sheetnames:
        del wb["metrics"]
    # place right after run_info (index 0) if present, else at front
    idx = 1 if "run_info" in wb.sheetnames else 0
    ws = wb.create_sheet("metrics", idx)

    r = 1
    ws.cell(r, 1, "Thinking-ablation quality metrics (latest run per job)").font = TITLE_FONT
    r += 2
    ws.cell(r, 1, f"reasoning_tags = {tags}").font = KEY_FONT
    r += 2

    # --- per (model, arm) ----------------------------------------------------
    hdr = ["model", "arm", "n", "reasoning_emitted_%", "truncated_%",
           "answer_isolated_%", "extract=score_%", "accuracy"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r, j, h); c.fill = SECTION_FILL; c.font = SECTION_FONT; c.alignment = CENTER
    r += 1
    groups: dict[tuple[str, str], list[vt.Diag]] = defaultdict(list)
    for d in diags:
        groups[(d.model, d.arm)].append(d)
    for model in MODELS:
        for arm in ARMS:
            rows = groups.get((model, arm), [])
            if not rows:
                continue
            n = len(rows)
            reason = _pct(sum(x.reasoning_present for x in rows), n) if arm == "think" else None
            trunc = _pct(sum(x.no_close_tag for x in rows), n) if arm == "think" else 0.0
            iso = _pct(sum(x.answer_isolated for x in rows), n)
            ext = _pct(sum(x.consistent for x in rows), n)
            acc = round(sum(x.score for x in rows) / n, 3)
            vals = [model, arm, n,
                    "—" if reason is None else reason, trunc, iso, ext, acc]
            for j, v in enumerate(vals, 1):
                cc = ws.cell(r, j, v)
                cc.alignment = LEFT if j <= 2 else CENTER
            r += 1
    r += 2

    # --- per-job token usage -------------------------------------------------
    ws.cell(r, 1, "Generated tokens per response, per job (latest run)").font = KEY_FONT
    r += 1
    hdr2 = ["job", "arm", "n", "tok_median", "tok_p90", "tok_max", "truncated_%"]
    for j, h in enumerate(hdr2, 1):
        c = ws.cell(r, j, h); c.fill = SECTION_FILL; c.font = SECTION_FONT; c.alignment = CENTER
    r += 1
    byjob: dict[str, list[vt.Diag]] = defaultdict(list)
    for d in diags:
        byjob[d.label].append(d)
    for label in sorted(byjob):
        rows = byjob[label]
        arm = rows[0].arm
        toks = [x.out_tokens for x in rows if x.out_tokens is not None]
        n = len(rows)
        trunc = _pct(sum(x.no_close_tag for x in rows), n) if arm == "think" else 0.0
        short = label
        for pfx in ("graph_bench_think_",):
            if short.startswith(pfx):
                short = short[len(pfx):]
        vals = [short, arm, n,
                _pctl(toks, 0.5), _pctl(toks, 0.9), max(toks) if toks else 0, trunc]
        for j, v in enumerate(vals, 1):
            cc = ws.cell(r, j, v)
            cc.alignment = LEFT if j == 1 else CENTER
        r += 1

    ws.cell(r + 1, 1,
            "reasoning_emitted = raw has a closed reasoning block; truncated = think arm, never "
            "closed (answer lost); answer_isolated = filtered text carries no reasoning tag; "
            "extract=score = re-normalizing filtered reproduces the logged score.").font = Font(italic=True)

    widths = {"A": 46, "B": 10, "C": 7, "D": 20, "E": 14, "F": 16, "G": 15, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def main() -> int:
    if len(sys.argv) != 2:
        print(f"usage: {sys.argv[0]} <xlsx_path>", file=sys.stderr)
        return 2
    path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(path)
    build_metrics(wb)
    wb.save(path)
    print(f"[thinking] added metrics -> {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
